"""XlsxRenderer — deterministic Excel generation from XlsxSpec."""
from __future__ import annotations

import io
import logging
import re
from typing import Optional

from app.rendering.doc_spec import XlsxSpec, XlsxSheetSpec

logger = logging.getLogger(__name__)

# P1-E: Excel built-in function whitelist for formula validation.
# Covers the ~100 most common functions; LLM-generated names outside this set are rejected.
_EXCEL_FUNCTIONS = {
    # Math & trig
    "SUM", "SUMIF", "SUMIFS", "SUMPRODUCT", "ABS", "ROUND", "ROUNDUP", "ROUNDDOWN",
    "INT", "MOD", "POWER", "SQRT", "EXP", "LN", "LOG", "LOG10", "FLOOR", "CEILING",
    "RAND", "RANDBETWEEN", "PI", "DEGREES", "RADIANS",
    # Stats
    "AVERAGE", "AVERAGEIF", "AVERAGEIFS", "COUNT", "COUNTA", "COUNTBLANK",
    "COUNTIF", "COUNTIFS", "MAX", "MIN", "MEDIAN", "MODE", "STDEV", "STDEVP",
    "VAR", "VARP", "LARGE", "SMALL", "RANK", "PERCENTILE", "QUARTILE",
    # Lookup
    "VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "LOOKUP", "XLOOKUP",
    "OFFSET", "INDIRECT", "CHOOSE", "ADDRESS",
    # Text
    "LEFT", "RIGHT", "MID", "LEN", "FIND", "SEARCH", "SUBSTITUTE", "REPLACE",
    "TRIM", "CLEAN", "UPPER", "LOWER", "PROPER", "CONCATENATE", "CONCAT",
    "TEXTJOIN", "TEXT", "VALUE", "NUMBERVALUE", "CHAR", "CODE",
    # Date & time
    "TODAY", "NOW", "DATE", "YEAR", "MONTH", "DAY", "HOUR", "MINUTE", "SECOND",
    "WEEKDAY", "WEEKNUM", "EOMONTH", "EDATE", "DATEDIF", "DAYS", "NETWORKDAYS",
    # Logical
    "IF", "IFS", "AND", "OR", "NOT", "IFERROR", "IFNA", "SWITCH", "XOR",
    # Financial
    "PMT", "FV", "PV", "RATE", "NPER", "NPV", "IRR", "XNPV", "XIRR",
    # Other
    "ISNUMBER", "ISTEXT", "ISBLANK", "ISERROR", "ISNA", "TYPE",
    "ROW", "COLUMN", "ROWS", "COLUMNS", "TRANSPOSE",
    # Dynamic array functions (Excel 365 / 2021+)
    "FILTER", "SORT", "SORTBY", "UNIQUE", "SEQUENCE", "RANDARRAY",
    "LET", "LAMBDA", "XLOOKUP", "XMATCH", "TOCOL", "TOROW",
    "VSTACK", "HSTACK", "TAKE", "DROP", "CHOOSEROWS", "CHOOSECOLS",
    "WRAPCOLS", "WRAPROWS", "EXPAND",
}


class XlsxRenderer:
    """Renders an XlsxSpec to .xlsx bytes."""

    def render(self, spec: XlsxSpec) -> bytes:
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for xlsx rendering") from exc

        wb = openpyxl.Workbook()
        if wb.active:
            wb.remove(wb.active)

        # P2-D: Set workbook metadata properties for proper document identity.
        import datetime as _dt
        wb.properties.title = spec.title or ""
        wb.properties.creator = "DataAgent Studio"
        wb.properties.lastModifiedBy = "DataAgent Studio"
        wb.properties.created = _dt.datetime.utcnow()
        wb.properties.modified = _dt.datetime.utcnow()

        # P1-5: Build spec_name → truncated_title map BEFORE filling sheets.
        # Sheet names are truncated to 31 chars by Excel; formulas that reference
        # the full spec.name will fail validation unless we track the mapping.
        truncated_name_map: dict[str, str] = {}
        for sheet_spec in spec.sheets:
            truncated = sheet_spec.name[:31]
            truncated_name_map[sheet_spec.name] = truncated

        for sheet_spec in spec.sheets:
            ws = wb.create_sheet(title=truncated_name_map[sheet_spec.name])
            self._fill_sheet(wb, ws, sheet_spec, name_map=truncated_name_map)

        # P1-D: Prepend auto-generated Dashboard/Summary sheet at position 0
        if len(spec.sheets) > 1:
            _insert_dashboard_sheet(wb, spec)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _fill_sheet(self, wb, ws, spec: XlsxSheetSpec, name_map: dict | None = None) -> None:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        row = 1
        table_header_row: Optional[int] = None
        table_col_count: int = 0
        n_data_rows: int = 0
        n_cols: int = 0

        # Description row
        if spec.description:
            ws.cell(row=row, column=1, value=spec.description)
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 2

        # Data table
        if spec.table and spec.table.headers:
            table_header_row = row
            n_cols = len(spec.table.headers)
            table_col_count = n_cols
            n_data_rows = len(spec.table.rows)

            for c_idx, h in enumerate(spec.table.headers, start=1):
                cell = ws.cell(row=row, column=c_idx, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="2563EB")
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for data_row in spec.table.rows:
                for c_idx, val in enumerate(data_row, start=1):
                    cell = ws.cell(row=row, column=c_idx, value=_coerce_value(val))
                    col_letter = get_column_letter(c_idx)
                    fmt = spec.number_formats.get(col_letter)
                    if fmt:
                        cell.number_format = fmt
                row += 1

            # P2-4: Auto-detect percentage columns (≥50% values end in '%') and
            # apply "0.0%" number format when no explicit format was given.
            _auto_apply_percent_formats(
                ws, table_header_row + 1, n_data_rows, n_cols, spec.number_formats
            )

            # P1-4: Register as a proper Excel Table object (enables AutoFilter +
            # structured references + proper pivot source)
            _add_excel_table(ws, spec.name, table_header_row, n_cols, n_data_rows)

            # P2-4: Add dropdown data validation for low-cardinality string columns
            if n_data_rows >= 2:
                _add_column_data_validations(ws, table_header_row, n_cols, n_data_rows)

            row += 1  # blank separator

        # Formulas
        for formula_spec in spec.formulas:
            # P1-5: Rewrite formula references using truncated sheet names
            _write_formula(ws, formula_spec, name_map=name_map)

        # Key findings
        if spec.key_findings:
            ws.cell(row=row, column=1, value="主要发现").font = Font(bold=True)
            row += 1
            for finding in spec.key_findings:
                ws.cell(row=row, column=1, value=f"• {finding}")
                row += 1
            row += 1

        # Calculation notes
        if spec.calculation_notes:
            ws.cell(row=row, column=1, value="计算说明").font = Font(bold=True, italic=True, size=9)
            row += 1
            for note in spec.calculation_notes:
                cell = ws.cell(row=row, column=1, value=note)
                cell.font = Font(italic=True, size=9, color="666666")
                row += 1

        # Charts
        if spec.charts:
            chart_anchor_row = table_header_row or 1
            for chart_spec in spec.charts:
                real_range = _resolve_data_range(
                    chart_spec.data_range, table_header_row, table_col_count
                )
                self._add_chart(ws, chart_spec, chart_anchor_row, real_range, table_col_count)
                chart_anchor_row += 16

        # Conditional formatting
        if spec.conditional_formats:
            _apply_conditional_formats(
                ws, spec.conditional_formats,
                table_header_row=table_header_row,
                n_data_rows=n_data_rows,
            )

        # P2-D: Named ranges
        if spec.named_ranges:
            _apply_named_ranges(wb, ws, spec.named_ranges)

        # P2-C: PivotTable (openpyxl-based approximation)
        if spec.pivot_table:
            _write_pivot_table(ws, spec.pivot_table, table_header_row, table_col_count)

        # Freeze pane
        if spec.freeze_pane:
            try:
                ws.freeze_panes = spec.freeze_pane
            except Exception:
                pass

        # P2-6: Page setup for printing — landscape, fit to one page wide
        try:
            from openpyxl.utils import get_column_letter as _gcl
            max_data_col = max(table_col_count, 1) if table_col_count else 4
            ws.print_area = f"A1:{_gcl(max_data_col)}{max(row - 1, 2)}"
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
        except Exception:
            pass

        # P1-4: Auto-width columns with CJK-aware character counting.
        # CJK characters are ~2x wide as ASCII in proportional fonts; weighting
        # them as 1.8 produces column widths that avoid text truncation.
        for col in ws.columns:
            max_len = max((_cjk_aware_len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)

    def _add_chart(
        self, ws, chart_spec, anchor_row: int, data_range: str, table_col_count: int
    ) -> None:
        try:
            from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart
            from openpyxl.chart import Reference
            from openpyxl.chart.series import DataPoint
            from openpyxl.utils.cell import range_boundaries, get_column_letter

            chart_type = chart_spec.chart_type.lower()
            if chart_type == "bar":
                chart = BarChart(); chart.type = "bar"
            elif chart_type == "column":
                chart = BarChart(); chart.type = "col"
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            elif chart_type == "area":
                chart = AreaChart()
            elif chart_type == "scatter":
                chart = ScatterChart()
            else:
                chart = BarChart(); chart.type = "col"

            chart.title = chart_spec.title
            chart.style = 10

            try:
                min_col, min_row, max_col, max_row = range_boundaries(data_range)
            except Exception:
                logger.warning("[Xlsx] Invalid data_range '%s', skipping chart", data_range)
                return

            data = Reference(ws, min_col=min_col, min_row=min_row,
                             max_col=max_col, max_row=max_row)
            chart.add_data(data, titles_from_data=True)

            # P2-1: Configure axis labels, numFmt, and Y-axis floor
            if chart_type not in ("pie", "donut"):
                try:
                    # Y-axis: format numbers and start from 0
                    chart.y_axis.numFmt = "#,##0"
                    chart.y_axis.scaling.min = 0
                    # Axis titles from chart spec title
                    title_parts = chart_spec.title.split("（")
                    if len(title_parts) > 1:
                        chart.y_axis.title = title_parts[1].rstrip("）") or ""
                    # X-axis: set category labels from the row above data
                    if min_row > 1:
                        cats = Reference(ws, min_col=min_col, max_col=max_col,
                                         min_row=min_row - 1, max_row=min_row - 1)
                        chart.set_categories(cats)
                except Exception:
                    pass

            anchor_col = max(table_col_count + 2, max_col + 2)
            anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
            ws.add_chart(chart, anchor)
        except Exception as exc:
            logger.warning("[Xlsx] Chart '%s' failed: %s", chart_spec.title, exc)


# ── Module-level helpers ──────────────────────────────────────────────────────

def _cjk_aware_len(text: str) -> float:
    """P1-4: Return display-width estimate for Excel column sizing.

    CJK unified ideographs, full-width punctuation, and Hangul/Kana are
    ~1.8× wider than ASCII in proportional fonts (e.g. 宋体/Calibri mixed).
    """
    import unicodedata
    width = 0.0
    for ch in text:
        cat = unicodedata.east_asian_width(ch)
        width += 1.8 if cat in ("W", "F") else 1.0
    return width


def _coerce_value(val):
    """Coerce cell values to Python numbers/dates where possible.

    P2-E: Try dateutil before falling back to string, so date-like values
    get written as datetime objects (Excel can then apply date number formats).
    """
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return val
    import datetime as _dt
    if isinstance(val, (_dt.date, _dt.datetime)):
        return val
    s = str(val).strip()

    # Strip numeric suffixes before trying number conversion
    for suffix in ("%", "万", "亿", "元", "人", "个", "次", "倍"):
        if s.endswith(suffix):
            try:
                return float(s[:-len(suffix)])
            except ValueError:
                break
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass

    # P2-E: Date detection — try common patterns first (fast), then dateutil
    _DATE_PATTERNS = [
        r"^\d{4}-\d{2}-\d{2}$",        # 2024-01-15
        r"^\d{4}/\d{2}/\d{2}$",         # 2024/01/15
        r"^\d{4}年\d{1,2}月\d{1,2}日$", # 2024年1月15日
        r"^\d{4}年\d{1,2}月$",          # 2024年1月
        r"^\d{4}-\d{2}$",               # 2024-01
    ]
    for pat in _DATE_PATTERNS:
        if re.match(pat, s):
            try:
                from dateutil import parser as _dateparser
                return _dateparser.parse(s.replace("年", "-").replace("月", "-").replace("日", ""))
            except Exception:
                break

    return s


def _write_formula(ws, formula_spec, name_map: dict | None = None) -> None:
    """Write a formula cell after validation: = prefix + balanced parens + function whitelist.

    P1-5: name_map (spec_name → truncated_title) is used to rewrite any cross-sheet
    references that use the full spec name so they match the actual worksheet title.
    """
    import re
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    formula = (formula_spec.formula or "").strip()
    if not formula.startswith("="):
        logger.warning("[Xlsx] Formula '%s' skipped: must start with '='", formula)
        return
    if formula.count("(") != formula.count(")"):
        logger.warning("[Xlsx] Formula '%s' skipped: unbalanced parentheses", formula)
        return

    if "!" in formula:
        # P1-5: Rewrite long spec names → truncated worksheet titles in the formula
        if name_map:
            for long_name, short_name in name_map.items():
                if long_name != short_name:
                    # Handle both quoted and unquoted refs
                    formula = formula.replace(f"'{long_name}'!", f"'{short_name}'!")
                    formula = formula.replace(f"{long_name}!", f"'{short_name}'!")

        # Validate that any sheet names referenced actually exist in the workbook
        existing_titles = {sh.title for sh in ws.parent.worksheets}
        for raw_sheet in re.findall(r"'([^']+)'!", formula) + re.findall(r"([A-Za-z一-鿿]\w*)!", formula):
            if raw_sheet not in existing_titles:
                logger.warning(
                    "[Xlsx] Formula '%s' skipped: sheet '%s' not found in workbook",
                    formula, raw_sheet,
                )
                return
    else:
        # P1-E: Only check function whitelist for pure in-sheet formulas.
        func_match = re.match(r"^=([A-Z][A-Z0-9_]*)\s*\(", formula.upper())
        if func_match:
            func_name = func_match.group(1)
            if func_name not in _EXCEL_FUNCTIONS:
                logger.warning("[Xlsx] Formula '%s' skipped: unknown function '%s'", formula, func_name)
                return

    try:
        col_letter, row_num = coordinate_from_string(formula_spec.cell)
        col_idx = column_index_from_string(col_letter)
        if formula_spec.label and col_idx > 1:
            import openpyxl
            label_cell = ws.cell(row=row_num, column=col_idx - 1, value=formula_spec.label)
            label_cell.font = openpyxl.styles.Font(bold=True, size=9)
        ws.cell(row=row_num, column=col_idx, value=formula)
    except Exception as exc:
        logger.warning("[Xlsx] Formula write failed (%s): %s", formula_spec.cell, exc)


def _resolve_data_range(
    llm_range: str,
    table_header_row: Optional[int],
    table_col_count: int,
) -> str:
    """Correct LLM-supplied data_range to match actual table position in the sheet.

    P1-3: Also handles whole-column references (A:A, A:C) that openpyxl cannot
    parse with range_boundaries() — these are converted to bounded cell ranges
    using the actual table extent so chart/formula references are valid.
    """
    if not table_header_row or not llm_range:
        return llm_range or "A1:C5"
    try:
        from openpyxl.utils.cell import range_boundaries, get_column_letter, column_index_from_string
        import re as _re_range

        # P2-4: Strip Excel absolute-reference $ signs — "$A$1:$C$5" → "A1:C5"
        llm_range = llm_range.replace("$", "")

        # Detect whole-column reference: "A:A" or "A:C" (letters only, no digits)
        _COL_RANGE_PAT = _re_range.compile(r'^([A-Z]+):([A-Z]+)$', _re_range.IGNORECASE)
        m = _COL_RANGE_PAT.match(llm_range.strip())
        if m:
            min_col_idx = column_index_from_string(m.group(1))
            max_col_idx = column_index_from_string(m.group(2))
            n_rows = max(table_col_count, 1)
            return (
                f"{get_column_letter(min_col_idx)}{table_header_row}"
                f":{get_column_letter(max_col_idx)}{table_header_row + n_rows}"
            )

        min_col, min_row, max_col, max_row = range_boundaries(llm_range)
        n_rows = max(max_row - min_row, 1)
        real_max_col = max(table_col_count, max_col - min_col + 1)
        real_max_row = table_header_row + n_rows
        return (
            f"{get_column_letter(1)}{table_header_row}"
            f":{get_column_letter(real_max_col)}{real_max_row}"
        )
    except Exception:
        return llm_range


def _apply_conditional_formats(
    ws,
    formats: list,
    table_header_row: Optional[int] = None,
    n_data_rows: int = 0,
) -> None:
    """Apply conditional formatting rules from XlsxConditionalFormat specs.

    P2-3: Clamp the LLM-supplied range so it never extends beyond the actual
    data rows — applying formats to empty rows wastes file space and can confuse
    Excel's print-area heuristics.
    """
    try:
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
        from openpyxl.styles import PatternFill
        from openpyxl.utils.cell import range_boundaries, get_column_letter

        for fmt in formats:
            cell_range = fmt.range if hasattr(fmt, "range") else fmt.get("range", "")
            rule_type = fmt.type if hasattr(fmt, "type") else fmt.get("type", "color_scale")
            threshold = fmt.threshold if hasattr(fmt, "threshold") else fmt.get("threshold", 0)

            if not cell_range:
                continue

            # P1-3: Pre-validate range syntax — openpyxl will crash with an
            # unhelpful error if the range string is malformed.  Skip silently.
            try:
                range_boundaries(cell_range)
            except Exception:
                logger.debug("[Xlsx] Skipping conditional format with invalid range '%s'", cell_range)
                continue

            # P2-3: Clamp range to actual data extent when we know it
            if table_header_row and n_data_rows > 0:
                try:
                    min_col, min_row, max_col, _ = range_boundaries(cell_range)
                    # Data rows begin at header_row + 1; clamp max row accordingly
                    data_last_row = table_header_row + n_data_rows
                    actual_min_row = max(min_row, table_header_row + 1)
                    actual_max_row = min(data_last_row, table_header_row + n_data_rows)
                    if actual_min_row > actual_max_row:
                        continue  # no data rows in range
                    cell_range = (
                        f"{get_column_letter(min_col)}{actual_min_row}"
                        f":{get_column_letter(max_col)}{actual_max_row}"
                    )
                except Exception:
                    pass  # keep original range on parse error

            if rule_type == "color_scale":
                rule = ColorScaleRule(
                    start_type="min",  start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max",    end_color="63BE7B",
                )
                ws.conditional_formatting.add(cell_range, rule)
            elif rule_type == "data_bar":
                rule = DataBarRule(
                    start_type="min", start_value=0,
                    end_type="max",   end_value=100,
                    color="638EC6",
                )
                ws.conditional_formatting.add(cell_range, rule)
            elif rule_type == "above_average":
                fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
                rule = CellIsRule(operator="greaterThan", formula=[str(threshold)], fill=fill)
                ws.conditional_formatting.add(cell_range, rule)

    except Exception as exc:
        logger.warning("[Xlsx] Conditional formatting failed: %s", exc)


def _auto_apply_percent_formats(
    ws,
    data_start_row: int,
    n_data_rows: int,
    n_cols: int,
    explicit_formats: dict,
) -> None:
    """P2-4: Detect columns whose values look like percentages and apply '0.0%' format.

    A column is considered a percentage column when ≥50% of its non-empty cell
    values are numeric strings ending in '%' (written by the LLM as e.g. '35%').
    The raw string is converted to a float (35% → 0.35) and the cell format set
    to '0.0%' so Excel renders it properly.
    """
    if n_data_rows == 0:
        return
    from openpyxl.utils.cell import get_column_letter

    for c_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(c_idx)
        if col_letter in explicit_formats:
            continue  # respect explicit override

        pct_count = 0
        total_count = 0
        pct_cells = []

        for r in range(data_start_row, data_start_row + n_data_rows):
            cell = ws.cell(row=r, column=c_idx)
            val = cell.value
            if val is None or val == "":
                continue
            total_count += 1
            if isinstance(val, str) and val.strip().endswith("%"):
                pct_count += 1
                pct_cells.append(cell)

        if total_count > 0 and pct_count / total_count >= 0.5:
            # Convert string percentages to float fractions and apply format
            for cell in pct_cells:
                try:
                    numeric = float(cell.value.strip().rstrip("%")) / 100
                    cell.value = numeric
                    cell.number_format = "0.0%"
                except Exception:
                    pass


def _apply_named_ranges(wb, ws, named_ranges: dict[str, str]) -> None:
    """P2-D: Register workbook-level named ranges."""
    for name, range_ref in named_ranges.items():
        try:
            # Qualify with sheet name: 'SheetName'!A1:C5
            sheet_name = ws.title
            qualified = f"'{sheet_name}'!{range_ref}" if "!" not in range_ref else range_ref
            wb.defined_names[name] = qualified
        except Exception as exc:
            logger.warning("[Xlsx] Named range '%s'='%s' failed: %s", name, range_ref, exc)


def _add_excel_table(ws, sheet_name: str, header_row: int, n_cols: int, n_data_rows: int) -> None:
    """P1-4: Wrap written data in a proper Excel Table object.

    This enables AutoFilter, banded-row styling, structured references like
    Table1[Revenue], and acts as a proper pivot source in Excel.
    """
    try:
        from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
        from openpyxl.utils import get_column_letter

        end_col = get_column_letter(n_cols)
        end_row = header_row + n_data_rows
        tbl_ref = f"A{header_row}:{end_col}{end_row}"
        # Sanitise name: Excel table names must start with a letter, no spaces/special chars
        safe_name = re.sub(r"[^a-zA-Z0-9_]", "_", sheet_name[:20]).lstrip("_") or "DataTable"
        xl_table = XlTable(displayName=safe_name, ref=tbl_ref)
        xl_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(xl_table)
    except Exception as exc:
        logger.warning("[Xlsx] Excel Table object creation failed: %s", exc)


def _add_column_data_validations(ws, header_row: int, n_cols: int, n_data_rows: int) -> None:
    """P2-4: Add dropdown data validation for low-cardinality string columns.
    P1-4: Also add decimal range validation (0–1) for ratio/percentage columns.

    Detects columns where all values are strings and there are 2–8 unique values,
    then adds an in-cell dropdown so users cannot enter arbitrary text.
    For columns with ratio/percentage headers (占比/比例/率/proportion/rate), adds
    a 0.0–1.0 decimal constraint so Excel rejects out-of-range entries.
    """
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter

        data_start = header_row + 1
        data_end = header_row + n_data_rows

        # P1-4: Keywords that indicate a ratio/proportion column (values should be 0–1)
        _RATIO_KEYWORDS = ("占比", "比例", "率", "proportion", "ratio", "rate", "pct", "percent")

        for c_idx in range(1, n_cols + 1):
            col_letter = get_column_letter(c_idx)
            header_val = str(ws.cell(row=header_row, column=c_idx).value or "").strip()
            header_lower = header_val.lower()

            # P1-4: Ratio column — add decimal 0–1 validation
            is_ratio_col = any(kw in header_lower for kw in _RATIO_KEYWORDS)
            if is_ratio_col:
                dv_ratio = DataValidation(
                    type="decimal",
                    operator="between",
                    formula1="0",
                    formula2="1",
                    allow_blank=True,
                    showErrorMessage=True,
                    error=f"'{header_val}' 应为 0–1 之间的小数（如 0.35 表示 35%）",
                    errorTitle="数值范围错误",
                )
                dv_ratio.add(f"{col_letter}{data_start}:{col_letter}{data_end}")
                ws.add_data_validation(dv_ratio)
                continue  # skip dropdown detection for ratio cols

            col_values = [
                ws.cell(row=r, column=c_idx).value
                for r in range(data_start, data_end + 1)
            ]
            str_values = [v for v in col_values if isinstance(v, str) and v.strip()]
            # Only add validation for all-string columns with 2–8 distinct values
            if len(str_values) != n_data_rows or not (2 <= len(set(str_values)) <= 8):
                continue
            unique = sorted(set(str_values))
            formula = '"' + ",".join(unique) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                                showErrorMessage=False)
            dv.add(f"{col_letter}{data_start}:{col_letter}{data_end}")
            ws.add_data_validation(dv)
    except Exception as exc:
        logger.warning("[Xlsx] Data validation setup failed: %s", exc)


def _write_pivot_table(ws, pivot_spec, table_header_row: Optional[int],
                       table_col_count: int) -> None:
    """P2-C: Write a pivot table approximation using openpyxl.

    openpyxl does not natively support live PivotTables (they require OOXML cache).
    We write a summary table instead — aggregated by the first row_field — which
    Excel can open and re-refresh as a pivot if needed.
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        if not table_header_row or not pivot_spec.row_fields or not pivot_spec.value_fields:
            logger.info("[Xlsx] Pivot skipped: insufficient spec (need row_fields + value_fields)")
            return

        # Collect actual header row from sheet
        headers = [
            ws.cell(row=table_header_row, column=c).value
            for c in range(1, table_col_count + 1)
        ]
        headers = [str(h) if h else "" for h in headers]

        # Find column indices for row_field and value_fields
        row_field = pivot_spec.row_fields[0]
        try:
            row_col_idx = headers.index(row_field) + 1
        except ValueError:
            logger.warning("[Xlsx] Pivot row_field '%s' not found in headers %s", row_field, headers)
            return

        val_col_indices = []
        for vf in pivot_spec.value_fields:
            try:
                val_col_indices.append((vf, headers.index(vf) + 1))
            except ValueError:
                logger.warning("[Xlsx] Pivot value_field '%s' not in headers, skipped", vf)

        if not val_col_indices:
            return

        # Collect data rows
        data_start = table_header_row + 1
        aggregated: dict[str, dict[str, float]] = {}
        for r in ws.iter_rows(min_row=data_start, values_only=True):
            row_key = str(r[row_col_idx - 1] or "（空）")
            if row_key not in aggregated:
                aggregated[row_key] = {vf: 0.0 for vf, _ in val_col_indices}
            for vf, c_idx in val_col_indices:
                raw = r[c_idx - 1]
                try:
                    aggregated[row_key][vf] = aggregated[row_key][vf] + float(raw or 0)
                except (TypeError, ValueError):
                    pass

        # Write pivot summary at dest_cell
        dest_col_letter, dest_row = coordinate_from_string(pivot_spec.dest_cell)
        dest_col = column_index_from_string(dest_col_letter)

        # Header row
        ws.cell(row=dest_row, column=dest_col, value=row_field).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=dest_row, column=dest_col).fill = PatternFill(fill_type="solid", fgColor="404040")
        for i, (vf, _) in enumerate(val_col_indices):
            c = dest_col + 1 + i
            ws.cell(row=dest_row, column=c, value=vf).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=dest_row, column=c).fill = PatternFill(fill_type="solid", fgColor="404040")

        # Data rows
        for offset, (key, vals) in enumerate(aggregated.items(), start=1):
            ws.cell(row=dest_row + offset, column=dest_col, value=key)
            for i, (vf, _) in enumerate(val_col_indices):
                ws.cell(row=dest_row + offset, column=dest_col + 1 + i, value=round(vals[vf], 2))

        logger.info("[Xlsx] Pivot summary written at %s (%d rows)", pivot_spec.dest_cell, len(aggregated))

    except Exception as exc:
        logger.warning("[Xlsx] Pivot table generation failed: %s", exc)


def _insert_dashboard_sheet(wb, spec) -> None:
    """P1-D: Prepend a '汇总' dashboard sheet at position 0 that cross-references
    all data sheets' key metrics for an at-a-glance summary.

    Strategy:
    - Reads the header row and first data row of each sheet's table
    - Writes a two-column summary (指标, 数值) using direct cell references
    - Inserts the sheet at index 0 so it is the first visible tab
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        dash = wb.create_sheet(title="汇总", index=0)
        row = 1

        # Title
        title_cell = dash.cell(row=row, column=1, value=f"{spec.title} — 数据汇总")
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        dash.merge_cells(f"A{row}:D{row}")
        row += 2

        for sheet_spec in spec.sheets:
            if not sheet_spec.table or not sheet_spec.table.headers:
                continue

            sheet_name = sheet_spec.name[:31]
            headers = sheet_spec.table.headers
            rows = sheet_spec.table.rows

            # P3-2: Section label — include sheet description when available
            label_text = f"▸ {sheet_name}"
            if sheet_spec.description:
                # Truncate description to keep the label concise (≤ 40 chars total)
                max_desc = max(10, 40 - len(sheet_name))
                desc_short = sheet_spec.description[:max_desc].rstrip()
                label_text = f"▸ {sheet_name}  —  {desc_short}"
            label_cell = dash.cell(row=row, column=1, value=label_text)
            label_cell.font = Font(bold=True, size=11)
            label_cell.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
            dash.merge_cells(f"A{row}:D{row}")
            row += 1

            # P1-4: Compute header/data row from spec structure — reliable,
            # no worksheet scanning required.
            # _fill_sheet() layout: [description row + blank] → header at row 3,
            # otherwise header at row 1.
            header_row_num = 3 if sheet_spec.description else 1
            data_row_num = header_row_num + 1

            for c_idx, header in enumerate(headers[:6], start=1):
                col_letter = get_column_letter(c_idx)
                # Label column
                dash.cell(row=row, column=1, value=str(header)).font = Font(bold=True, size=9)
                # P1-6: Use reliable direct cell reference (universally compatible).
                # Structured Table refs have cross-sheet compatibility issues in some
                # Excel versions; direct refs are simpler and always work.
                ref = f"='{sheet_name}'!{col_letter}{data_row_num}"
                ref_cell = dash.cell(row=row, column=2, value=ref)
                ref_cell.alignment = Alignment(horizontal="right")

                # Sparkline: show trend across all data rows for this column
                # (requires ≥3 data rows to be meaningful)
                n_data = len(sheet_spec.table.rows) if sheet_spec.table.rows else 0
                if n_data >= 3:
                    spark_range = f"'{sheet_name}'!{col_letter}{data_row_num}:{col_letter}{data_row_num + n_data - 1}"
                    spark_cell = f"C{row}"
                    _add_sparkline(dash, spark_range, spark_cell)

                row += 1

            row += 1  # blank line between sheets

        # Auto-width columns A and B
        dash.column_dimensions["A"].width = 30
        dash.column_dimensions["B"].width = 20

        # Column C header
        dash.cell(row=1, column=3, value="趋势").font = Font(bold=True, size=9, color="666666")
        dash.column_dimensions["C"].width = 15

        logger.info("[Xlsx] Dashboard sheet created (%d rows)", row)
    except Exception as exc:
        logger.warning("[Xlsx] Dashboard sheet creation failed: %s", exc)


def _add_sparkline(ws, data_range: str, target_cell: str) -> None:
    """Add a line sparkline to target_cell sourcing data from data_range.

    Uses openpyxl's SparklineGroup API (available in openpyxl ≥3.1).
    P2-6: When the API is unavailable, writes an ASCII trend string computed
    from the actual data values so the cell is always informative (not blank).
    E.g. "10 ▲ 15 ▲ 12 ▼ 18" showing direction changes across data points.
    """
    try:
        from openpyxl.chart.sparkline import Sparkline, SparklineGroup
        sparkline = Sparkline()
        sparkline.ref = data_range
        sparkline.sqref = target_cell
        grp = SparklineGroup(ref=data_range, sqref=target_cell, type="line")
        grp.sparklines.append(sparkline)
        if not hasattr(ws, "sparkline_groups"):
            _write_ascii_trend(ws, data_range, target_cell)
            return
        ws.sparkline_groups.append(grp)
    except Exception as exc:
        logger.debug("[Xlsx] Sparkline API failed (%s), using ASCII trend", exc)
        _write_ascii_trend(ws, data_range, target_cell)


def _write_ascii_trend(ws, data_range: str, target_cell: str) -> None:
    """P2-6: Write an ASCII trend indicator extracted from the source data range.

    Reads up to 5 numeric values from the data_range cells, formats them as
    'val1 ▲/▼ val2 ▲/▼ ...' showing direction of change between each pair.
    Falls back to '→ (数据)' when values cannot be read.
    """
    try:
        from openpyxl.utils.cell import range_boundaries
        from openpyxl.styles import Font as _Font

        # Parse data range (handles cross-sheet refs like 'Sheet1'!A2:A10)
        clean_range = data_range.split("!")[-1].strip("'\"")
        min_col, min_row, max_col, max_row = range_boundaries(clean_range)

        # Read values (column-major for single-column ranges)
        values = []
        for r in range(min_row, min(max_row + 1, min_row + 5)):
            for c in range(min_col, max_col + 1):
                cell_val = ws.cell(row=r, column=c).value
                try:
                    values.append(float(cell_val or 0))
                except (TypeError, ValueError):
                    pass

        if not values:
            trend_str = "→ (数据)"
        elif len(values) == 1:
            trend_str = str(int(values[0]) if values[0] == int(values[0]) else round(values[0], 1))
        else:
            parts = []
            for i, v in enumerate(values[:5]):
                label = str(int(v) if v == int(v) else round(v, 1))
                if i > 0:
                    arrow = "▲" if v > values[i - 1] else ("▼" if v < values[i - 1] else "─")
                    parts.append(f"{arrow}{label}")
                else:
                    parts.append(label)
            trend_str = " ".join(parts)

        cell = ws[target_cell]
        cell.value = trend_str
        cell.font = _Font(size=7, color="444444")
    except Exception as exc2:
        logger.debug("[Xlsx] ASCII trend fallback failed: %s", exc2)
