"""Structured data service — DuckDB-based SQL path for Excel/CSV uploads.

Why this exists:
  RAG vector search is designed for unstructured text.  When users upload Excel or
  CSV files with rows/columns (e.g., branch performance tables), vector search loses
  the relational structure — numbers become fuzzy text chunks.  This service provides
  a parallel "structured data" path:

    upload Excel/CSV  →  load into DuckDB in-memory DB  →  describe schema to LLM
    →  LLM generates SQL  →  execute precisely  →  clean tabular results for reports

Usage:
  sds = StructuredDataService()
  await sds.ingest_file(kb_id, doc_id, file_path, file_name)
  schema = sds.describe_tables(kb_id)              # → str summary for LLM context
  results = sds.execute_query(kb_id, sql)           # → list[dict]
  # or let the LLM generate SQL and run it:
  results = await sds.nl_query(kb_id, question, llm_context)
"""
from __future__ import annotations

import io
import json
import logging
import os
import re
import threading
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

# Thread-local DuckDB connections (DuckDB connections are not thread-safe to share)
_conn_cache: dict[str, Any] = {}
_lock = threading.Lock()


def _get_conn(kb_id: str):
    """Return a per-KB DuckDB in-memory connection (created on first access)."""
    import duckdb
    with _lock:
        if kb_id not in _conn_cache:
            # Use a file-backed DB so data survives process restarts
            db_dir = Path(os.getenv("DATA_DIR", "/app/data")) / "structured_dbs"
            db_dir.mkdir(parents=True, exist_ok=True)
            db_path = str(db_dir / f"{kb_id}.duckdb")
            _conn_cache[kb_id] = duckdb.connect(db_path)
        return _conn_cache[kb_id]


def _safe_table_name(name: str) -> str:
    """Convert a filename to a valid SQL table name."""
    stem = Path(name).stem
    safe = re.sub(r"[^\w]", "_", stem)
    safe = re.sub(r"_+", "_", safe).strip("_")
    if not safe or safe[0].isdigit():
        safe = "t_" + safe
    return safe[:60]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class StructuredDataService:
    """Manages DuckDB tables for structured file uploads per knowledge base."""

    def ingest_file(
        self,
        kb_id: str,
        doc_id: str,
        file_path: str,
        file_name: str,
    ) -> dict:
        """Load an Excel/CSV file into DuckDB and return table metadata.

        Returns:
            {"table_name": str, "row_count": int, "columns": [{"name": str, "type": str}]}
        """
        conn = _get_conn(kb_id)
        table_name = _safe_table_name(file_name)
        suffix = Path(file_name).suffix.lower()

        try:
            if suffix in {".xlsx", ".xls", ".xlsm"}:
                df = _read_excel(file_path)
            elif suffix in {".csv", ".tsv"}:
                sep = "\t" if suffix == ".tsv" else ","
                import pandas as pd
                # Smart encoding detection (handles Chinese GBK/GB18030 from Excel-Windows)
                try:
                    from app.services.file_parser import detect_encoding
                    with open(file_path, "rb") as fb:
                        encoding = detect_encoding(fb.read(65536))
                except Exception:
                    encoding = "utf-8"
                df = pd.read_csv(
                    file_path,
                    sep=sep,
                    encoding=encoding,
                    encoding_errors="replace",
                    low_memory=False,
                )
                df = _infer_column_types(df)
            else:
                raise ValueError(f"Unsupported file type: {suffix}")

            df.columns = [_safe_col(c) for c in df.columns]

            # Register as DuckDB table (replace if exists)
            conn.execute(f"DROP TABLE IF EXISTS \"{table_name}\"")
            conn.register("__tmp_df__", df)
            conn.execute(f'CREATE TABLE "{table_name}" AS SELECT * FROM __tmp_df__')
            conn.unregister("__tmp_df__")

            # Record metadata in a registry table
            _ensure_registry(conn)
            conn.execute(
                "INSERT OR REPLACE INTO __registry__ (table_name, doc_id, file_name, row_count) VALUES (?,?,?,?)",
                [table_name, doc_id, file_name, len(df)],
            )

            columns = _describe_columns(conn, table_name)
            logger.info("Ingested %s → table '%s' (%d rows, %d cols)", file_name, table_name, len(df), len(columns))
            return {"table_name": table_name, "row_count": len(df), "columns": columns}

        except Exception as exc:
            logger.error("Failed to ingest %s into DuckDB: %s", file_name, exc, exc_info=True)
            raise

    def list_tables(self, kb_id: str) -> list[dict]:
        """Return all tables registered for a KB."""
        conn = _get_conn(kb_id)
        _ensure_registry(conn)
        rows = conn.execute("SELECT table_name, doc_id, file_name, row_count FROM __registry__").fetchall()
        return [{"table_name": r[0], "doc_id": r[1], "file_name": r[2], "row_count": r[3]} for r in rows]

    def describe_tables(self, kb_id: str, max_sample_rows: int = 3) -> str:
        """Generate a compact schema description string for injection into LLM prompts.

        Format:
            表名: branch_data  (行数: 1523)
            列: 分行名称(TEXT), 贷款余额(DOUBLE), 存款余额(DOUBLE), 净利润(DOUBLE), ...
            样本数据 (前3行):
            | 分行名称 | 贷款余额 | ...
            | 华东分行 | 4523.2  | ...
        """
        tables = self.list_tables(kb_id)
        if not tables:
            return ""
        conn = _get_conn(kb_id)
        parts = []
        for t in tables:
            name = t["table_name"]
            cols = _describe_columns(conn, name)
            col_str = ", ".join(f"{c['name']}({c['type']})" for c in cols[:20])
            try:
                sample = conn.execute(f'SELECT * FROM "{name}" LIMIT {max_sample_rows}').fetchdf()
                sample_md = sample.to_markdown(index=False)
            except Exception:
                sample_md = ""
            part = (
                f"【结构化数据表】{name}  (共 {t['row_count']} 行)\n"
                f"字段: {col_str}\n"
                f"样本:\n{sample_md}"
            )
            parts.append(part)
        return "\n\n".join(parts)

    def execute_query(
        self,
        kb_id: str,
        sql: str,
        max_rows: int = 500,
        *,
        question: str = "",
    ) -> list[dict]:
        """Execute a SQL query and return rows as list of dicts.

        Only SELECT statements are permitted.  Raises ValueError for mutating SQL.
        On success, records the (question, sql) pair in the few-shot cache.
        """
        stripped = sql.strip().lstrip(";").strip()
        first_word = stripped.split()[0].upper() if stripped else ""
        if first_word not in {"SELECT", "WITH", "EXPLAIN"}:
            raise ValueError(f"Only SELECT queries are permitted (got {first_word!r})")

        conn = _get_conn(kb_id)
        try:
            df = conn.execute(sql).fetchdf()
            row_count = len(df)
            if row_count > max_rows:
                df = df.head(max_rows)
            results = df.to_dict(orient="records")

            # Cache the successful query for future few-shot retrieval
            if question and row_count > 0:
                _ensure_registry(conn)
                _record_query_success(conn, question, sql, row_count)

            return results
        except Exception as exc:
            logger.warning("SQL query failed: %s\nSQL: %s", exc, sql)
            raise

    def get_few_shot_examples(self, kb_id: str, question: str, top_k: int = 3) -> list[dict]:
        """Retrieve top-k similar past successful queries for the given question.

        Returned as list of {"question": str, "sql": str, "row_count": int}.
        Empty list if no similar queries cached yet.
        """
        conn = _get_conn(kb_id)
        _ensure_registry(conn)
        return _retrieve_similar_queries(conn, question, top_k=top_k)

    def get_table_relationships(self, kb_id: str) -> list[dict]:
        """Return inferred JOIN relationships between this KB's tables.

        Used by the SQL prompt to enable multi-table queries the LLM might
        otherwise hesitate to construct.
        """
        conn = _get_conn(kb_id)
        return infer_table_relationships(conn)

    def get_sample_queries(self, kb_id: str) -> list[str]:
        """Auto-generate example SQL queries for each table — useful for LLM few-shot context."""
        tables = self.list_tables(kb_id)
        queries = []
        conn = _get_conn(kb_id)
        for t in tables[:4]:
            name = t["table_name"]
            cols = _describe_columns(conn, name)
            num_cols = [c["name"] for c in cols if c["type"] in {"DOUBLE", "FLOAT", "INTEGER", "BIGINT", "HUGEINT"}]
            text_cols = [c["name"] for c in cols if c["type"] in {"VARCHAR", "TEXT"}]

            if num_cols and text_cols:
                agg_cols = ", ".join(f'SUM("{c}")' for c in num_cols[:3])
                q = f'SELECT "{text_cols[0]}", {agg_cols} FROM "{name}" GROUP BY "{text_cols[0]}" ORDER BY 2 DESC LIMIT 20'
                queries.append(q)
            elif num_cols:
                add_cols = "+".join(f'"{c}"' for c in num_cols[:2])
                q = f'SELECT *, ({add_cols}) AS total FROM "{name}" ORDER BY total DESC LIMIT 20'
                queries.append(q)
            else:
                queries.append(f'SELECT * FROM "{name}" LIMIT 10')
        return queries

    def drop_table(self, kb_id: str, table_name: str) -> None:
        """Remove a table (called when document is deleted from KB)."""
        conn = _get_conn(kb_id)
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        conn.execute("DELETE FROM __registry__ WHERE table_name = ?", [table_name])

    async def nl_query(
        self,
        kb_id: str,
        question: str,
        *,
        max_retries: int = 2,
    ) -> dict:
        """自然语言 → SQL → 执行，返回结果字典（与 DuckDBEngine 接口对齐）."""
        from app.pipeline.llm_helpers import call_llm_json

        schema_text = self.describe_tables(kb_id)
        if not schema_text:
            return {"success": False, "error": "该知识库没有结构化数据表，请先上传 CSV / Excel 文件。", "sql": ""}

        # Include table relationships for better multi-table queries
        relationships = self.get_table_relationships(kb_id)
        rel_text = ""
        if relationships:
            rel_text = "\n表间关联:\n" + "\n".join(
                f"  {r['table_a']}.{r['col_a']} ↔ {r['table_b']}.{r['col_b']} ({r['reason']})"
                for r in relationships[:6]
            )

        # Include few-shot examples
        few_shot = self.get_few_shot_examples(kb_id)
        few_shot_text = ""
        if few_shot:
            few_shot_text = "\n参考查询:\n" + "\n".join(
                f"  Q: {q}\n  SQL: {sql}"
                for q, sql in few_shot[:3]
            )

        system_prompt = f"""你是专业的 SQL 分析师。根据用户问题和数据库 Schema 生成 DuckDB SQL 查询。
仅输出 JSON: {{"sql": "SELECT ..."}}
不要添加任何解释。
SQL 方言为 DuckDB，支持 PIVOT/UNPIVOT、窗口函数、正则、日期函数。

可用数据表 Schema:
{schema_text}{rel_text}{few_shot_text}"""

        user_msg = f"问题: {question}"
        last_error: str | None = None
        generated_sql = ""

        for attempt in range(max_retries + 1):
            if attempt > 0 and last_error:
                user_msg = f"上次 SQL 报错: {last_error}\n请修正 SQL 并重新生成。\n问题: {question}"

            try:
                resp = await call_llm_json(system_prompt, user_msg)
                sql = resp.get("sql", "").strip()
                if not sql:
                    last_error = "LLM 返回空 SQL"
                    continue
                generated_sql = sql
                result = self.execute_query(kb_id, sql, question=question)
                return {
                    "success": True,
                    "sql": sql,
                    "rows": result,
                    "row_count": len(result),
                }
            except ValueError as ve:
                last_error = str(ve)
                # Non-SELECT queries are rejected; don't retry
                return {"success": False, "error": last_error, "sql": generated_sql}
            except Exception as exc:
                last_error = str(exc)
                logger.warning("NL→SQL attempt %d failed for kb=%s: %s", attempt, kb_id, exc)

        return {"success": False, "error": f"SQL 生成失败: {last_error}", "sql": generated_sql}

    def close(self, kb_id: str) -> None:
        """Close and remove the DuckDB connection for a KB."""
        with _lock:
            conn = _conn_cache.pop(kb_id, None)
        if conn:
            try:
                conn.close()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_col(name: str) -> str:
    """Sanitize a DataFrame column name."""
    if not isinstance(name, str):
        name = str(name)
    # Strip leading/trailing whitespace and special chars
    name = name.strip()
    return name if name else "col"


def _ensure_registry(conn) -> None:
    conn.execute(
        """CREATE TABLE IF NOT EXISTS __registry__ (
            table_name VARCHAR PRIMARY KEY,
            doc_id     VARCHAR,
            file_name  VARCHAR,
            row_count  BIGINT,
            created_at TIMESTAMP DEFAULT current_timestamp
        )"""
    )
    # Few-shot query cache (Vanna.ai-inspired): remember successful NL→SQL pairs
    # so future queries on the same schema can use them as examples
    conn.execute(
        """CREATE TABLE IF NOT EXISTS __query_cache__ (
            id           INTEGER PRIMARY KEY,
            question     VARCHAR,
            sql_text     VARCHAR,
            tables_used  VARCHAR,
            row_count    BIGINT,
            executed_at  TIMESTAMP DEFAULT current_timestamp,
            success      BOOLEAN DEFAULT true
        )"""
    )
    # Auto-increment for the cache
    conn.execute("CREATE SEQUENCE IF NOT EXISTS __qc_seq__ START 1")


def _record_query_success(conn, question: str, sql_text: str, row_count: int) -> None:
    """Persist a successful (question, sql) pair for future few-shot retrieval."""
    if not question or not sql_text:
        return
    try:
        # Extract referenced table names heuristically
        tables = re.findall(r'FROM\s+"?(\w+)"?|JOIN\s+"?(\w+)"?', sql_text, re.IGNORECASE)
        tables_used = ",".join({t for pair in tables for t in pair if t})

        conn.execute(
            """INSERT INTO __query_cache__ (id, question, sql_text, tables_used, row_count, success)
               VALUES (nextval('__qc_seq__'), ?, ?, ?, ?, true)""",
            [question[:500], sql_text[:2000], tables_used[:500], row_count],
        )
        # Cap cache size at 200 most recent
        conn.execute(
            """DELETE FROM __query_cache__ WHERE id NOT IN (
                SELECT id FROM __query_cache__ ORDER BY executed_at DESC LIMIT 200
            )"""
        )
    except Exception as exc:
        logger.debug("Query cache record failed: %s", exc)


def _retrieve_similar_queries(conn, question: str, top_k: int = 3) -> list[dict]:
    """Retrieve top-k semantically similar past successful queries.

    Uses lightweight token-overlap scoring — no embedding required, fast,
    and surprisingly effective for the small (<200) cache size.
    """
    if not question:
        return []
    try:
        rows = conn.execute(
            "SELECT question, sql_text, row_count FROM __query_cache__ WHERE success = true"
        ).fetchall()
    except Exception:
        return []

    if not rows:
        return []

    # Token overlap scoring (Chinese bigram + ASCII word)
    q_tokens = _query_tokenize(question)
    if not q_tokens:
        return []

    scored: list[tuple[float, dict]] = []
    for q_text, sql_text, row_count in rows:
        cand_tokens = _query_tokenize(q_text or "")
        if not cand_tokens:
            continue
        overlap = len(q_tokens & cand_tokens) / max(len(q_tokens), 1)
        if overlap < 0.15:
            continue
        scored.append((overlap, {
            "question": q_text,
            "sql": sql_text,
            "row_count": int(row_count or 0),
        }))

    scored.sort(key=lambda x: x[0], reverse=True)
    return [r for _, r in scored[:top_k]]


def _query_tokenize(text: str) -> set[str]:
    """Lightweight tokenization for query similarity (bigrams + ASCII words)."""
    if not text:
        return set()
    text = text.lower()
    tokens: set[str] = set()
    # ASCII words
    for w in re.findall(r"[a-z]{3,}", text):
        tokens.add(w)
    # Chinese bigrams
    for run in re.findall(r"[一-龥]+", text):
        for i in range(len(run) - 1):
            tokens.add(run[i:i + 2])
    return tokens


def _describe_columns(conn, table_name: str) -> list[dict]:
    """Return [{name, type}] for a DuckDB table."""
    try:
        rows = conn.execute(f"DESCRIBE \"{table_name}\"").fetchall()
        return [{"name": r[0], "type": r[1]} for r in rows]
    except Exception:
        return []


def _read_excel(file_path: str):
    """Read an Excel file to a pandas DataFrame.

    SOTA improvements over baseline:
      • Detect merged cells via openpyxl + forward-fill them so data isn't lost
      • Detect multi-row headers (when row 1 + row 2 form a hierarchy) and flatten
      • Type inference for date / currency / percentage columns
      • Multi-sheet: preserve sheet metadata as both __sheet__ AND __sheet_order__
        columns (the latter enables temporal sorting when sheets are year-named)
    """
    import pandas as pd
    xf = pd.ExcelFile(file_path)
    sheet_names = xf.sheet_names[:8]

    dfs: list = []
    for order, sheet in enumerate(sheet_names):
        df = _read_one_sheet(file_path, sheet)
        if df is None or df.empty:
            continue
        if len(sheet_names) > 1:
            # Preserve sheet metadata (lossless context for downstream SQL)
            df.insert(0, "__sheet__", str(sheet))
            df.insert(1, "__sheet_order__", order)
        dfs.append(df)

    if not dfs:
        return pd.DataFrame()
    if len(dfs) == 1:
        return dfs[0]

    # Concat: outer join so non-overlapping columns are preserved as NaN
    return pd.concat(dfs, ignore_index=True, sort=False)


def _read_one_sheet(file_path: str, sheet_name):
    """Read one sheet with merged-cell forward-fill + multi-row-header detection."""
    import pandas as pd

    # ── Step 1: detect merged cell ranges via openpyxl ──────────────────────
    merged_ranges = _detect_merged_ranges(file_path, sheet_name)

    # ── Step 2: detect multi-row headers heuristically ──────────────────────
    header_rows = _detect_header_depth(file_path, sheet_name)

    try:
        if header_rows > 1:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=list(range(header_rows)))
            # Flatten hierarchical columns: ('A', 'X') → 'A_X'
            df.columns = [
                "_".join(str(c).strip() for c in col if c and not str(c).startswith("Unnamed"))
                or f"col_{i}"
                for i, col in enumerate(df.columns)
            ]
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
    except Exception as exc:
        logger.warning("Sheet %s read failed (%s); falling back to header=0", sheet_name, exc)
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

    # ── Step 3: forward-fill merged-cell holes ──────────────────────────────
    if merged_ranges:
        df = _fill_merged_cells(df, merged_ranges, header_offset=header_rows)

    # ── Step 4: type inference for date / numeric / percentage columns ──────
    df = _infer_column_types(df)

    return df


def _detect_merged_ranges(file_path: str, sheet_name) -> list[tuple]:
    """Return list of (min_row, min_col, max_row, max_col) for merged cells.

    Rows/cols are 1-indexed (openpyxl convention). Returns empty list if not xlsx
    or no merges. Uses data_only=True so formula cells return their computed
    cached values (set by Excel last save) — never raw "=A1+B1" strings.

    IMPORTANT: data_only=True relies on Excel having cached the result. If a
    formula-heavy file was created programmatically (e.g. by xlsxwriter) and
    NEVER opened in Excel, the cache is empty and formula cells will return
    None. We detect this case and warn (caller can run a formula evaluator).
    """
    suffix = Path(file_path).suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        ranges = [(r.min_row, r.min_col, r.max_row, r.max_col) for r in ws.merged_cells.ranges]

        # Detect uncached-formulas problem: if the file declares formulas but
        # the cached values are mostly None, warn so caller can fall back.
        try:
            wb2 = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
            ws2 = wb2[sheet_name]
            has_formula = False
            for row in ws2.iter_rows(max_row=50, values_only=False):
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        has_formula = True
                        break
                if has_formula:
                    break
            wb2.close()
            if has_formula:
                # Check how many cached values are None in the data_only workbook
                none_count = sum(
                    1 for row in ws.iter_rows(max_row=50, values_only=True)
                    for v in row if v is None
                )
                total_cells = 50 * (ws.max_column or 1)
                if total_cells and none_count / total_cells > 0.6:
                    logger.warning(
                        "Sheet '%s' has formulas with uncached values "
                        "(%d/%d cells empty). Consider opening + saving in Excel.",
                        sheet_name, none_count, total_cells,
                    )
        except Exception:
            pass

        wb.close()
        return ranges
    except Exception as exc:
        logger.debug("Merged-cell detection failed for %s/%s: %s", file_path, sheet_name, exc)
        return []


# ─────────────────────────────────────────────────────────────────────────────
# Cross-table relationship inference (auto-detect potential JOIN keys)
# ─────────────────────────────────────────────────────────────────────────────

def infer_table_relationships(conn) -> list[dict]:
    """Detect potential JOIN relationships between tables in this KB.

    Heuristic: for each pair of (table_a, table_b), find column-name pairs
    where:
      • Names are identical or one is a clear suffix of the other (e.g.
        "branch_id" vs "id" in branch_data)
      • Column types match
      • At least one column has the structure of a foreign key
        (low cardinality / repeating values in one table)

    Returns: list of {table_a, col_a, table_b, col_b, confidence, reason}

    This is information the LLM can USE in its SQL prompt: "If asked to combine
    these tables, JOIN on (table_a.col_a = table_b.col_b)."
    """
    try:
        rows = conn.execute(
            "SELECT table_name FROM __registry__ WHERE table_name NOT LIKE '__%'"
        ).fetchall()
    except Exception:
        return []

    table_names = [r[0] for r in rows]
    if len(table_names) < 2:
        return []

    # Collect column metadata per table
    table_meta: dict[str, list[dict]] = {}
    for t in table_names:
        try:
            cols = conn.execute(f'DESCRIBE "{t}"').fetchall()
            table_meta[t] = [{"name": c[0], "type": c[1]} for c in cols]
        except Exception:
            continue

    relationships: list[dict] = []
    table_pairs = [
        (a, b) for i, a in enumerate(table_names) for b in table_names[i + 1:]
    ]

    for table_a, table_b in table_pairs:
        cols_a = table_meta.get(table_a, [])
        cols_b = table_meta.get(table_b, [])
        if not cols_a or not cols_b:
            continue

        for ca in cols_a:
            for cb in cols_b:
                # Type match
                if ca["type"] != cb["type"]:
                    continue
                # Skip wildcard sheet metadata
                if ca["name"].startswith("__") or cb["name"].startswith("__"):
                    continue
                # Name match
                conf, reason = _score_column_match(ca["name"], cb["name"], ca["type"])
                if conf < 0.5:
                    continue
                relationships.append({
                    "table_a": table_a,
                    "col_a": ca["name"],
                    "table_b": table_b,
                    "col_b": cb["name"],
                    "confidence": round(conf, 2),
                    "reason": reason,
                })

    # Sort by confidence
    relationships.sort(key=lambda r: -r["confidence"])
    return relationships[:8]


def _score_column_match(name_a: str, name_b: str, col_type: str) -> tuple[float, str]:
    """Score how likely two columns are to participate in a join."""
    a_lower = name_a.lower().strip()
    b_lower = name_b.lower().strip()

    if a_lower == b_lower:
        # Exact match
        if "id" in a_lower or a_lower in ("code", "key", "no", "编号", "代码"):
            return 0.95, "列名完全相同且为典型 ID 列"
        return 0.75, "列名完全相同"

    # Substring match (e.g. "branch" vs "branch_id")
    if a_lower in b_lower or b_lower in a_lower:
        shorter, longer = sorted([a_lower, b_lower], key=len)
        if longer.endswith("_" + shorter) or longer.endswith(shorter):
            return 0.7, f"\"{longer}\" 看起来是 \"{shorter}\" 的外键引用"
        return 0.55, "列名为子串关系"

    # Bigram overlap (Chinese-friendly)
    bigrams_a = {a_lower[i:i + 2] for i in range(len(a_lower) - 1)}
    bigrams_b = {b_lower[i:i + 2] for i in range(len(b_lower) - 1)}
    if bigrams_a and bigrams_b:
        overlap = len(bigrams_a & bigrams_b) / max(len(bigrams_a | bigrams_b), 1)
        if overlap >= 0.5:
            return 0.55 + overlap * 0.2, "列名词形高度相似"

    return 0.0, ""


def _detect_header_depth(file_path: str, sheet_name) -> int:
    """Heuristic: how many rows form the header?

    Looks at the first 4 rows. A row is "header-like" if it's mostly strings
    with no numeric content. We return the count of consecutive header-like
    rows starting from row 0. Capped at 3 (multi-level headers beyond 3 are rare).
    """
    import pandas as pd
    try:
        peek = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=4)
    except Exception:
        return 1

    if peek.empty or len(peek) < 2:
        return 1

    depth = 0
    for idx in range(min(3, len(peek) - 1)):
        row = peek.iloc[idx]
        # Header-like: >70% non-null + 0 numeric
        non_null = row.notna().mean()
        numeric_count = sum(1 for v in row if isinstance(v, (int, float)) and not pd.isna(v))
        if non_null >= 0.7 and numeric_count == 0:
            depth += 1
        else:
            break
    return max(1, depth)


def _fill_merged_cells(df, merged_ranges: list[tuple], header_offset: int = 1):
    """Forward-fill values that were merged in the original Excel.

    A merged cell in Excel has its value ONLY in the top-left cell; the rest
    are blank. When we read into a DataFrame, those become NaN. We use the
    merged_ranges metadata to restore the values, which is critical for
    downstream SQL (so GROUP BY on a merged category works).
    """
    if not merged_ranges or df.empty:
        return df

    # Convert merged ranges to (df_row_start, df_col_start, df_row_end, df_col_end)
    # openpyxl is 1-indexed; the first `header_offset` rows are consumed as header
    for min_row, min_col, max_row, max_col in merged_ranges:
        df_row_start = max_row - 1 - header_offset  # last header row → df row 0
        df_row_end = max_row - 1 - header_offset    # if range spans rows
        df_col = min_col - 1                         # 0-indexed
        if df_row_start < 0 or df_col < 0 or df_col >= df.shape[1]:
            continue
        # If the merge spans multiple data rows, forward-fill within column
        if max_row > min_row:
            try:
                top_val = df.iat[min_row - 1 - header_offset, df_col]
                if top_val is None:
                    continue
                for r in range(min_row - header_offset, max_row - header_offset):
                    if 0 <= r < df.shape[0]:
                        try:
                            cur = df.iat[r, df_col]
                            if cur is None or (hasattr(cur, "__len__") is False and pd.isna(cur)):
                                df.iat[r, df_col] = top_val
                        except Exception:
                            continue
            except Exception:
                continue
    return df


def _infer_column_types(df):
    """Promote columns to richer types: dates, percentages, currency.

    pandas's default inference often leaves these as `object`. We do one
    explicit pass per column:
      • If column name contains 日期/时间/date → try pd.to_datetime
      • If values contain % → strip % and convert to float (divide by 100)
      • If values contain ￥ or $ or ,1000 separator → strip and convert to float
    """
    import pandas as pd

    for col in df.columns:
        col_lower = str(col).lower()
        # Only process string-like columns; skip already-numeric / datetime
        if not pd.api.types.is_string_dtype(df[col]) and df[col].dtype != object:
            continue

        # Date columns
        if any(kw in str(col) for kw in ("日期", "时间", "年月", "Date", "Time")) or "date" in col_lower:
            try:
                converted = pd.to_datetime(df[col], errors="coerce")
                # Only accept if ≥60% successfully converted
                if converted.notna().mean() >= 0.6:
                    df[col] = converted
                    continue
            except Exception:
                pass

        # Percentage / currency / number-with-separator
        sample = df[col].dropna().astype(str).head(20)
        if len(sample) == 0:
            continue

        if (sample.str.contains("%").mean()) >= 0.5:
            try:
                stripped = df[col].astype(str).str.replace("%", "", regex=False).str.replace(",", "", regex=False)
                df[col] = pd.to_numeric(stripped, errors="coerce") / 100.0
                continue
            except Exception:
                pass

        if sample.str.contains(r"[￥$€£,]").mean() >= 0.4:
            try:
                stripped = df[col].astype(str).str.replace(r"[￥$€£,\s]", "", regex=True)
                df[col] = pd.to_numeric(stripped, errors="coerce")
                continue
            except Exception:
                pass

        # Plain numeric-as-string
        if sample.str.match(r"^-?\d+(\.\d+)?$").mean() >= 0.8:
            try:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            except Exception:
                pass

    return df


def is_structured_file(file_name: str) -> bool:
    """Return True if the file should go through the SQL path instead of RAG."""
    return Path(file_name).suffix.lower() in {".xlsx", ".xls", ".xlsm", ".csv", ".tsv"}


# Singleton
_instance: Optional[StructuredDataService] = None


def get_structured_data_service() -> StructuredDataService:
    global _instance
    if _instance is None:
        _instance = StructuredDataService()
    return _instance
