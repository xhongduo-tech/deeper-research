"""Offline workbook QA for generated Excel artifacts."""
from __future__ import annotations

import io
from typing import Any


def _score(items: list[dict[str, Any]]) -> float:
    if not items:
        return 1.0
    return sum(1 for item in items if item["passed"]) / len(items)


def score_xlsx_workbook(
    workbook: bytes | str,
    *,
    source_registry: list[dict] | None = None,
) -> dict:
    """Inspect a real XLSX artifact and return SOTA-style quality signals.

    The check is intentionally offline and deterministic. It focuses on the
    delivery properties that matter for data analysis reports: workbook
    structure, formulas, visual objects, data provenance, and analyst ergonomics.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment dependent
        return {
            "passed": False,
            "overall_score": 0,
            "error": f"openpyxl 不可用: {exc}",
            "warnings": ["无法读取真实 XLSX，Excel QA 已降级"],
            "blockers": ["EXCEL_QA_UNAVAILABLE"],
            "dimensions": {},
        }

    try:
        if isinstance(workbook, bytes):
            wb = openpyxl.load_workbook(io.BytesIO(workbook), data_only=False)
        else:
            wb = openpyxl.load_workbook(workbook, data_only=False)
    except Exception as exc:
        return {
            "passed": False,
            "overall_score": 0,
            "error": f"无法打开 XLSX: {exc}",
            "warnings": [],
            "blockers": ["EXCEL_UNREADABLE"],
            "dimensions": {},
        }

    source_registry = source_registry or []
    sheet_names = wb.sheetnames
    non_empty_sheets = []
    formula_count = 0
    chart_count = 0
    numeric_cells = 0
    source_note_count = 0
    frozen_panes_count = 0
    table_like_sheets = 0

    for ws in wb.worksheets:
        values = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if cell.value not in (None, "")
        ]
        if values:
            non_empty_sheets.append(ws.title)
        if ws.freeze_panes:
            frozen_panes_count += 1
        chart_count += len(getattr(ws, "_charts", []) or [])
        row_lengths = []
        for row in ws.iter_rows():
            row_values = [cell.value for cell in row if cell.value not in (None, "")]
            if row_values:
                row_lengths.append(len(row_values))
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                if isinstance(value, (int, float)):
                    numeric_cells += 1
                if isinstance(value, str) and any(marker in value for marker in ("来源", "口径", "假设", "数据范围", "清洗", "单位")):
                    source_note_count += 1
        if sum(1 for length in row_lengths if length >= 3) >= 2:
            table_like_sheets += 1

    descriptive_names = [
        name for name in sheet_names
        if not name.lower().startswith("sheet") and len(name.strip()) >= 2
    ]
    has_summary = bool(sheet_names) and any(
        marker in sheet_names[0] for marker in ("摘要", "总览", "Summary", "报告")
    )

    structure = [
        {"id": "XLSX-S001", "name": "readable_workbook", "passed": bool(sheet_names)},
        {"id": "XLSX-S002", "name": "has_summary_first_sheet", "passed": has_summary},
        {"id": "XLSX-S003", "name": "descriptive_sheet_names", "passed": len(descriptive_names) == len(sheet_names)},
        {"id": "XLSX-S004", "name": "non_empty_sheets", "passed": len(non_empty_sheets) == len(sheet_names)},
    ]
    analysis = [
        {"id": "XLSX-A001", "name": "table_like_outputs", "passed": table_like_sheets >= 1},
        {"id": "XLSX-A002", "name": "numeric_cells_present", "passed": numeric_cells > 0},
        {"id": "XLSX-A003", "name": "formulas_present", "passed": formula_count > 0},
        {"id": "XLSX-A004", "name": "charts_or_multiple_tables", "passed": chart_count > 0 or table_like_sheets >= 2},
    ]
    provenance = [
        {"id": "XLSX-P001", "name": "source_registry_or_notes", "passed": bool(source_registry) or source_note_count > 0},
        {"id": "XLSX-P002", "name": "frozen_panes_for_review", "passed": frozen_panes_count > 0},
    ]

    dimensions = {
        "structure": {"score": round(_score(structure) * 100, 1), "items": structure},
        "analysis": {"score": round(_score(analysis) * 100, 1), "items": analysis},
        "provenance": {"score": round(_score(provenance) * 100, 1), "items": provenance},
    }
    overall = round(sum(v["score"] for v in dimensions.values()) / len(dimensions), 1)
    blockers = [
        item["id"]
        for item in structure + analysis + provenance
        if not item["passed"] and item["id"] in {"XLSX-S001", "XLSX-S004", "XLSX-A001", "XLSX-P001"}
    ]
    warnings = [
        item["id"]
        for item in structure + analysis + provenance
        if not item["passed"] and item["id"] not in blockers
    ]

    return {
        "passed": not blockers and overall >= 72,
        "overall_score": overall,
        "blockers": blockers,
        "warnings": warnings,
        "sheet_count": len(sheet_names),
        "non_empty_sheet_count": len(non_empty_sheets),
        "formula_count": formula_count,
        "chart_count": chart_count,
        "numeric_cell_count": numeric_cells,
        "source_note_count": source_note_count,
        "frozen_panes_count": frozen_panes_count,
        "table_like_sheet_count": table_like_sheets,
        "dimensions": dimensions,
    }
